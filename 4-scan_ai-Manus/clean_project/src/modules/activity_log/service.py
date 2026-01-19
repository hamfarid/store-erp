"""
مسار الملف: /home/ubuntu/gaara_scan_ai_final_4.2/src/modules/activity_log/service.py
الوصف: خدمات مديول سجل النشاط
المؤلف: فريق Gaara ERP
تاريخ الإنشاء: 29 مايو 2025
"""

from typing import Dict, Any, List, Optional, Tuple, Union
from datetime import datetime, timedelta, timezone
import json
import csv
import io
import logging
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, or_
from fastapi import HTTPException, BackgroundTasks

from src.modules.activity_log import models, schemas
from src.modules.user_management.models import User

# إعداد التسجيل
logger = logging.getLogger(__name__)


async def create_activity_log(
    db: Session,
    log_data: Union[schemas.ActivityLogCreate, schemas.UserActionLogCreate, schemas.SystemActionLogCreate, schemas.AIActionLogCreate],
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    إنشاء سجل نشاط جديد

    Args:
        db (Session): جلسة قاعدة البيانات
        log_data (Union[ActivityLogCreate, UserActionLogCreate, SystemActionLogCreate, AIActionLogCreate]): بيانات سجل النشاط
        background_tasks (Optional[BackgroundTasks]): مهام خلفية

    Returns:
        models.ActivityLog: سجل النشاط المنشأ
    """
    try:
        # إنشاء سجل النشاط
        activity_log = models.ActivityLog(
            log_type=log_data.log_type,
            module_id=log_data.module_id,
            action_id=log_data.action_id,
            user_id=log_data.user_id,
            description=log_data.description,
            log_data=log_data.log_data,
            ip_address=log_data.ip_address,
            user_agent=log_data.user_agent
        )

        db.add(activity_log)
        db.commit()
        db.refresh(activity_log)

        # إذا كان هناك مهام خلفية، أضف مهمة لإنشاء السجل المتخصص
        if background_tasks:
            if log_data.log_type == "system":
                background_tasks.add_task(
                    create_system_log_from_activity,
                    db,
                    activity_log.id,
                    log_data.log_data
                )
            elif log_data.log_type == "user":
                background_tasks.add_task(
                    create_user_log_from_activity,
                    db,
                    activity_log.id,
                    log_data.log_data
                )
            elif log_data.log_type == "ai":
                background_tasks.add_task(
                    create_ai_log_from_activity,
                    db,
                    activity_log.id,
                    log_data.log_data
                )

        return activity_log

    except Exception as e:
        db.rollback()
        logger.error("خطأ في إنشاء سجل النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في إنشاء سجل النشاط: {str(e)}") from e


async def create_system_log_from_activity(
    db: Session,
    activity_log_id: int,
    details: Dict[str, Any]
) -> models.SystemLog:
    """
    إنشاء سجل نظام من سجل نشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط
        details (Dict[str, Any]): تفاصيل السجل

    Returns:
        models.SystemLog: سجل النظام المنشأ
    """
    try:
        # استخراج البيانات من التفاصيل
        component = details.get("component", "unknown")
        event_type = details.get("event_type", "unknown")
        severity = details.get("severity", "info")
        message = details.get("message", "")

        # إنشاء سجل النظام
        system_log = models.SystemLog(
            activity_log_id=activity_log_id,
            component=component,
            event_type=event_type,
            severity=severity,
            message=message,
            details=details
        )

        db.add(system_log)
        db.commit()
        db.refresh(system_log)

        return system_log

    except Exception as e:
        db.rollback()
        logger.error("خطأ في إنشاء سجل النظام: %s", str(e))
        return None


async def create_user_log_from_activity(
    db: Session,
    activity_log_id: int,
    details: Dict[str, Any]
) -> models.UserLog:
    """
    إنشاء سجل مستخدم من سجل نشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط
        details (Dict[str, Any]): تفاصيل السجل

    Returns:
        models.UserLog: سجل المستخدم المنشأ
    """
    try:
        # استخراج البيانات من التفاصيل
        user_id = details.get("user_id")
        session_id = details.get("session_id")
        action_type = details.get("action_type", "unknown")
        resource_type = details.get("resource_type")
        resource_id = details.get("resource_id")
        before_state = details.get("before_state")
        after_state = details.get("after_state")

        # إنشاء سجل المستخدم
        user_log = models.UserLog(
            activity_log_id=activity_log_id,
            user_id=user_id,
            session_id=session_id,
            action_type=action_type,
            resource_type=resource_type,
            resource_id=resource_id,
            before_state=before_state,
            after_state=after_state
        )

        db.add(user_log)
        db.commit()
        db.refresh(user_log)

        return user_log

    except Exception as e:
        db.rollback()
        logger.error("خطأ في إنشاء سجل المستخدم: %s", str(e))
        return None


async def create_ai_log_from_activity(
    db: Session,
    activity_log_id: int,
    details: Dict[str, Any]
) -> models.AILog:
    """
    إنشاء سجل ذكاء اصطناعي من سجل نشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط
        details (Dict[str, Any]): تفاصيل السجل

    Returns:
        models.AILog: سجل الذكاء الاصطناعي المنشأ
    """
    try:
        # استخراج البيانات من التفاصيل
        agent_id = details.get("agent_id", "unknown")
        agent_type = details.get("agent_type", "unknown")
        interaction_type = details.get("interaction_type", "query")
        query = details.get("query")
        response = details.get("response")
        tokens_used = details.get("tokens_used")
        processing_time = details.get("processing_time")
        confidence_score = details.get("confidence_score")

        # إنشاء سجل الذكاء الاصطناعي
        ai_log = models.AILog(
            activity_log_id=activity_log_id,
            agent_id=agent_id,
            agent_type=agent_type,
            interaction_type=interaction_type,
            query=query,
            response=response,
            tokens_used=tokens_used,
            processing_time=processing_time,
            confidence_score=confidence_score
        )

        db.add(ai_log)
        db.commit()
        db.refresh(ai_log)

        return ai_log

    except Exception as e:
        db.rollback()
        logger.error("خطأ في إنشاء سجل الذكاء الاصطناعي: %s", str(e))
        return None


async def get_activity_logs(
    db: Session,
    filters: schemas.ActivityLogFilter
) -> Tuple[List[models.ActivityLog], int]:
    """
    الحصول على سجلات النشاط مع التصفية

    Args:
        db (Session): جلسة قاعدة البيانات
        filters (schemas.ActivityLogFilter): معايير التصفية

    Returns:
        Tuple[List[models.ActivityLog], int]: سجلات النشاط وإجمالي العدد
    """
    try:
        # إنشاء الاستعلام الأساسي
        query = db.query(models.ActivityLog)

        # تطبيق التصفية
        if filters.log_type:
            query = query.filter(models.ActivityLog.log_type == filters.log_type)

        if filters.module_id:
            query = query.filter(models.ActivityLog.module_id == filters.module_id)

        if filters.action_id:
            query = query.filter(models.ActivityLog.action_id == filters.action_id)

        if filters.user_id:
            query = query.filter(models.ActivityLog.user_id == filters.user_id)

        if filters.status:
            query = query.filter(models.ActivityLog.status == filters.status)

        if filters.start_date:
            query = query.filter(models.ActivityLog.created_at >= filters.start_date)

        if filters.end_date:
            query = query.filter(models.ActivityLog.created_at <= filters.end_date)

        if filters.search:
            search_term = f"%{filters.search}%"
            query = query.filter(
                or_(
                    models.ActivityLog.description.ilike(search_term),
                    models.ActivityLog.module_id.ilike(search_term),
                    models.ActivityLog.action_id.ilike(search_term)
                )
            )

        # الحصول على إجمالي العدد
        total = query.count()

        # تطبيق الترتيب والتقسيم إلى صفحات
        query = query.order_by(desc(models.ActivityLog.created_at))
        query = query.offset((filters.page - 1) * filters.page_size).limit(filters.page_size)

        # تنفيذ الاستعلام
        logs = query.all()

        return logs, total

    except Exception as e:
        logger.error("خطأ في الحصول على سجلات النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على سجلات النشاط: {str(e)}") from e


async def get_activity_log_by_id(
    db: Session,
    log_id: int
) -> models.ActivityLog:
    """
    الحصول على سجل نشاط بواسطة المعرف

    Args:
        db (Session): جلسة قاعدة البيانات
        log_id (int): معرف سجل النشاط

    Returns:
        models.ActivityLog: سجل النشاط
    """
    try:
        log = db.query(models.ActivityLog).filter(models.ActivityLog.id == log_id).first()

        if not log:
            raise HTTPException(status_code=404, detail="سجل النشاط غير موجود")

        return log

    except Exception as e:
        logger.error("خطأ في الحصول على سجل النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على سجل النشاط: {str(e)}") from e


async def get_system_log_by_activity_id(
    db: Session,
    activity_log_id: int
) -> models.SystemLog:
    """
    الحصول على سجل النظام بواسطة معرف سجل النشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط

    Returns:
        models.SystemLog: سجل النظام
    """
    try:
        log = db.query(models.SystemLog).filter(models.SystemLog.activity_log_id == activity_log_id).first()

        if not log:
            raise HTTPException(status_code=404, detail="سجل النظام غير موجود")

        return log

    except Exception as e:
        logger.error("خطأ في الحصول على سجل النظام: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على سجل النظام: {str(e)}") from e


async def get_user_log_by_activity_id(
    db: Session,
    activity_log_id: int
) -> models.UserLog:
    """
    الحصول على سجل المستخدم بواسطة معرف سجل النشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط

    Returns:
        models.UserLog: سجل المستخدم
    """
    try:
        log = db.query(models.UserLog).filter(models.UserLog.activity_log_id == activity_log_id).first()

        if not log:
            raise HTTPException(status_code=404, detail="سجل المستخدم غير موجود")

        return log

    except Exception as e:
        logger.error("خطأ في الحصول على سجل المستخدم: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على سجل المستخدم: {str(e)}") from e


async def get_ai_log_by_activity_id(
    db: Session,
    activity_log_id: int
) -> models.AILog:
    """
    الحصول على سجل الذكاء الاصطناعي بواسطة معرف سجل النشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        activity_log_id (int): معرف سجل النشاط

    Returns:
        models.AILog: سجل الذكاء الاصطناعي
    """
    try:
        log = db.query(models.AILog).filter(models.AILog.activity_log_id == activity_log_id).first()

        if not log:
            raise HTTPException(status_code=404, detail="سجل الذكاء الاصطناعي غير موجود")

        return log

    except Exception as e:
        logger.error("خطأ في الحصول على سجل الذكاء الاصطناعي: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على سجل الذكاء الاصطناعي: {str(e)}") from e


async def get_log_modules(
    db: Session
) -> List[models.LogModule]:
    """
    الحصول على مديولات السجل

    Args:
        db (Session): جلسة قاعدة البيانات

    Returns:
        List[models.LogModule]: مديولات السجل
    """
    try:
        return db.query(models.LogModule).filter(models.LogModule.is_active == 1).all()

    except Exception as e:
        logger.error("خطأ في الحصول على مديولات السجل: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على مديولات السجل: {str(e)}") from e


async def get_log_actions(
    db: Session,
    module_id: Optional[str] = None
) -> List[models.LogAction]:
    """
    الحصول على إجراءات السجل

    Args:
        db (Session): جلسة قاعدة البيانات
        module_id (Optional[str]): معرف المديول

    Returns:
        List[models.LogAction]: إجراءات السجل
    """
    try:
        query = db.query(models.LogAction).filter(models.LogAction.is_active == 1)

        if module_id:
            query = query.filter(models.LogAction.module_id == module_id)

        return query.all()

    except Exception as e:
        logger.error("خطأ في الحصول على إجراءات السجل: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على إجراءات السجل: {str(e)}") from e


async def get_log_users(
    db: Session
) -> List[User]:
    """
    الحصول على مستخدمي السجل

    Args:
        db (Session): جلسة قاعدة البيانات

    Returns:
        List[User]: مستخدمي السجل
    """
    try:
        return db.query(User).all()

    except Exception as e:
        logger.error("خطأ في الحصول على مستخدمي السجل: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على مستخدمي السجل: {str(e)}") from e


async def get_activity_log_statistics(
    db: Session,
    days: int = 30
) -> schemas.ActivityLogStatistics:
    """
    الحصول على إحصائيات سجل النشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        days (int): عدد الأيام

    Returns:
        schemas.ActivityLogStatistics: إحصائيات سجل النشاط
    """
    try:
        # تحديد تاريخ البداية
        start_date = datetime.now(timezone.utc) - timedelta(days=days)

        # إجمالي السجلات
        total_logs = db.query(func.count(models.ActivityLog.id)).filter(  # pylint: disable=not-callable
            models.ActivityLog.created_at >= start_date
        ).scalar()

        # السجلات حسب النوع
        logs_by_type = {}
        for log_type in ["system", "user", "ai"]:
            count = db.query(func.count(models.ActivityLog.id)).filter(  # pylint: disable=not-callable
                models.ActivityLog.log_type == log_type,
                models.ActivityLog.created_at >= start_date
            ).scalar()
            logs_by_type[log_type] = count

        # السجلات حسب المديول
        logs_by_module = {}
        module_counts = db.query(
            models.ActivityLog.module_id,
            func.count(models.ActivityLog.id)  # pylint: disable=not-callable
        ).filter(
            models.ActivityLog.created_at >= start_date
        ).group_by(
            models.ActivityLog.module_id
        ).all()

        for module_id, count in module_counts:
            logs_by_module[module_id] = count

        # السجلات حسب الحالة
        logs_by_status = {}
        status_counts = db.query(
            models.ActivityLog.status,
            func.count(models.ActivityLog.id)  # pylint: disable=not-callable
        ).filter(
            models.ActivityLog.created_at >= start_date
        ).group_by(
            models.ActivityLog.status
        ).all()

        for status, count in status_counts:
            logs_by_status[status] = count

        # السجلات حسب التاريخ
        logs_by_date = {}
        for i in range(days):
            date = datetime.now(timezone.utc) - timedelta(days=i)
            date_str = date.strftime("%Y-%m-%d")

            count = db.query(func.count(models.ActivityLog.id)).filter(  # pylint: disable=not-callable
                func.date(models.ActivityLog.created_at) == func.date(date)
            ).scalar()

            logs_by_date[date_str] = count

        return schemas.ActivityLogStatistics(
            total_logs=total_logs,
            logs_by_type=logs_by_type,
            logs_by_module=logs_by_module,
            logs_by_status=logs_by_status,
            logs_by_date=logs_by_date
        )

    except Exception as e:
        logger.error("خطأ في الحصول على إحصائيات سجل النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في الحصول على إحصائيات سجل النشاط: {str(e)}") from e


async def export_activity_logs(
    db: Session,
    export_data: schemas.ActivityLogExport
) -> bytes:
    """
    تصدير سجلات النشاط

    Args:
        db (Session): جلسة قاعدة البيانات
        export_data (schemas.ActivityLogExport): بيانات التصدير

    Returns:
        bytes: بيانات التصدير
    """
    try:
        # الحصول على السجلات
        logs, _ = await get_activity_logs(db, export_data.filter)

        # تحويل السجلات إلى قائمة من القواميس
        log_dicts = []
        for log in logs:
            log_dict = {
                "id": log.id,
                "log_type": log.log_type,
                "module_id": log.module_id,
                "action_id": log.action_id,
                "user_id": log.user_id,
                "description": log.description,
                "status": log.status,
                "created_at": log.created_at.isoformat()
            }
            log_dicts.append(log_dict)

        # تصدير حسب الصيغة
        if export_data.export_format == "csv":
            return export_to_csv(log_dicts)
        elif export_data.export_format == "xlsx":
            return export_to_xlsx(log_dicts)
        elif export_data.export_format == "pdf":
            return export_to_pdf(log_dicts)
        else:
            raise HTTPException(status_code=400, detail="صيغة التصدير غير مدعومة")

    except HTTPException:
        raise

    except Exception as e:
        logger.error("خطأ في تصدير سجلات النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تصدير سجلات النشاط: {str(e)}") from e


def export_to_csv(logs: List[Dict[str, Any]]) -> bytes:
    """
    تصدير السجلات إلى CSV

    Args:
        logs (List[Dict[str, Any]]): السجلات

    Returns:
        bytes: بيانات CSV
    """
    try:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=logs[0].keys())
        writer.writeheader()
        writer.writerows(logs)

        return output.getvalue().encode("utf-8")

    except Exception as e:
        logger.error("خطأ في تصدير السجلات إلى CSV: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تصدير السجلات إلى CSV: {str(e)}") from e


def export_to_xlsx(logs: List[Dict[str, Any]]) -> bytes:
    """
    تصدير السجلات إلى XLSX

    Args:
        logs (List[Dict[str, Any]]): السجلات

    Returns:
        bytes: بيانات XLSX
    """
    try:
        # استخدام مكتبة openpyxl لإنشاء ملف Excel
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activity Logs"

        # كتابة العناوين
        headers = list(logs[0].keys())
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)

        # كتابة البيانات
        for row_num, log in enumerate(logs, 2):
            for col_num, key in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = log[key]

        # حفظ الملف إلى بايتس
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    except Exception as e:
        logger.error("خطأ في تصدير السجلات إلى XLSX: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تصدير السجلات إلى XLSX: {str(e)}") from e


def export_to_pdf(logs: List[Dict[str, Any]]) -> bytes:
    """
    تصدير السجلات إلى PDF

    Args:
        logs (List[Dict[str, Any]]): السجلات

    Returns:
        bytes: بيانات PDF
    """
    try:
        # استخدام مكتبة reportlab لإنشاء ملف PDF
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        # إنشاء ملف PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []

        # إضافة العنوان
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Activity Logs", styles["Title"]))

        # إنشاء جدول البيانات
        headers = list(logs[0].keys())
        data = [headers]
        for log in logs:
            row = [str(log[key]) for key in headers]
            data.append(row)

        # إنشاء الجدول
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

        # بناء الملف
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        logger.error("خطأ في تصدير السجلات إلى PDF: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تصدير السجلات إلى PDF: {str(e)}") from e


async def cleanup_old_logs(
    db: Session,
    background_tasks: BackgroundTasks
) -> Dict[str, Any]:
    """
    تنظيف السجلات القديمة

    Args:
        db (Session): جلسة قاعدة البيانات
        background_tasks (BackgroundTasks): مهام خلفية

    Returns:
        Dict[str, Any]: نتيجة التنظيف
    """
    try:
        # الحصول على سياسات الاحتفاظ
        retention_policies = db.query(models.LogRetentionPolicy).all()

        # إضافة مهمة خلفية لكل نوع سجل
        for policy in retention_policies:
            background_tasks.add_task(
                cleanup_logs_by_type,
                db,
                policy.log_type,
                policy.retention_days,
                policy.archive_enabled,
                policy.archive_path
            )

        return {"status": "success", "message": "تم جدولة تنظيف السجلات القديمة"}

    except Exception as e:
        logger.error("خطأ في تنظيف السجلات القديمة: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تنظيف السجلات القديمة: {str(e)}") from e


async def cleanup_logs_by_type(
    db: Session,
    log_type: str,
    retention_days: int,
    archive_enabled: bool,
    archive_path: Optional[str]
) -> None:
    """
    تنظيف السجلات حسب النوع

    Args:
        db (Session): جلسة قاعدة البيانات
        log_type (str): نوع السجل
        retention_days (int): عدد أيام الاحتفاظ
        archive_enabled (bool): تفعيل الأرشفة
        archive_path (Optional[str]): مسار الأرشيف
    """
    try:
        # تحديد تاريخ القطع
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=retention_days)

        # الحصول على السجلات القديمة
        old_logs = db.query(models.ActivityLog).filter(
            models.ActivityLog.log_type == log_type,
            models.ActivityLog.created_at < cutoff_date
        ).all()

        # أرشفة السجلات إذا كان مطلوباً
        if archive_enabled and archive_path and old_logs:
            await archive_logs(old_logs, log_type, archive_path)

        # حذف السجلات القديمة
        db.query(models.ActivityLog).filter(
            models.ActivityLog.log_type == log_type,
            models.ActivityLog.created_at < cutoff_date
        ).delete(synchronize_session=False)

        db.commit()

        logger.info("تم تنظيف %d سجل من نوع %s", len(old_logs), log_type)

    except Exception as e:
        db.rollback()
        logger.error("خطأ في تنظيف السجلات من نوع %s: %s", log_type, str(e))


async def archive_logs(
    logs: List[models.ActivityLog],
    log_type: str,
    archive_path: str
) -> None:
    """
    أرشفة السجلات

    Args:
        logs (List[models.ActivityLog]): السجلات
        log_type (str): نوع السجل
        archive_path (str): مسار الأرشيف
    """
    try:
        import os

        # إنشاء مسار الأرشيف إذا لم يكن موجوداً
        os.makedirs(archive_path, exist_ok=True)

        # إنشاء اسم ملف الأرشيف
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        archive_file = os.path.join(archive_path, f"{log_type}_logs_{timestamp}.json")

        # تحويل السجلات إلى JSON
        log_dicts = []
        for log in logs:
            log_dict = {
                "id": log.id,
                "log_type": log.log_type,
                "module_id": log.module_id,
                "action_id": log.action_id,
                "user_id": log.user_id,
                "description": log.description,
                "details": log.details,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "status": log.status,
                "created_at": log.created_at.isoformat()
            }
            log_dicts.append(log_dict)

        # كتابة السجلات إلى ملف الأرشيف
        with open(archive_file, "w", encoding="utf-8") as f:
            json.dump(log_dicts, f, ensure_ascii=False, indent=2)

        logger.info("تم أرشفة %d سجل من نوع %s إلى %s", len(logs), log_type, archive_file)

    except Exception as e:
        logger.error("خطأ في أرشفة السجلات من نوع %s: %s", log_type, str(e))


# دوال مساعدة للتكامل مع المديولات الأخرى

async def log_system_event(
    db: Session,
    component: str,
    event_type: str,
    severity: str,
    message: str,
    details: Optional[Dict[str, Any]] = None,
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    تسجيل حدث نظام

    Args:
        db (Session): جلسة قاعدة البيانات
        component (str): المكون
        event_type (str): نوع الحدث
        severity (str): مستوى الخطورة
        message (str): الرسالة
        details (Optional[Dict[str, Any]]): تفاصيل إضافية
        background_tasks (Optional[BackgroundTasks]): مهام خلفية

    Returns:
        models.ActivityLog: سجل النشاط المنشأ
    """
    # تحديد حالة السجل بناءً على مستوى الخطورة
    status_map = {
        "critical": "error",
        "error": "error",
        "warning": "warning",
        "info": "info",
        "debug": "info"
    }

    status = status_map.get(severity, "info")

    # إنشاء تفاصيل السجل
    log_details = {
        "component": component,
        "event_type": event_type,
        "severity": severity,
        "message": message
    }

    if details:
        log_details.update(details)

    # إنشاء سجل النشاط
    log_data = schemas.ActivityLogCreate(
        log_type="system",
        module_id=component,
        action_id=event_type,
        user_id=None,
        description=message,
        details=log_details,
        status=status
    )

    return await create_activity_log(db, log_data, background_tasks)


async def log_user_action(
    db: Session,
    action_data: Union[schemas.UserActionLogCreate, schemas.ActivityLogCreate],
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    تسجيل إجراء مستخدم

    Args:
        db (Session): جلسة قاعدة البيانات
        action_data (Union[UserActionLogCreate, ActivityLogCreate]): بيانات إجراء المستخدم
        background_tasks (Optional[BackgroundTasks]): مهام خلفية

    Returns:
        models.ActivityLog: سجل النشاط
    """
    try:
        # إنشاء سجل النشاط
        log_data = schemas.ActivityLogCreate(
            log_type="user",
            module_id=action_data.module_id,
            action_id=action_data.action_id,
            user_id=action_data.user_id,
            description=action_data.description,
            log_data={
                "user_id": action_data.user_id,
                "action_type": action_data.action_type if hasattr(action_data, 'action_type') else None,
                "resource_type": action_data.resource_type if hasattr(action_data, 'resource_type') else None,
                "resource_id": action_data.resource_id if hasattr(action_data, 'resource_id') else None,
                "before_state": action_data.before_state if hasattr(action_data, 'before_state') else None,
                "after_state": action_data.after_state if hasattr(action_data, 'after_state') else None
            },
            ip_address=action_data.ip_address,
            user_agent=action_data.user_agent
        )

        return await create_activity_log(db, log_data, background_tasks)

    except Exception as e:
        logger.error("خطأ في تسجيل إجراء المستخدم: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تسجيل إجراء المستخدم: {str(e)}") from e


async def log_ai_interaction(
    db: Session,
    interaction_data: schemas.AIInteractionLogCreate,
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    تسجيل تفاعل ذكاء اصطناعي

    Args:
        db (Session): جلسة قاعدة البيانات
        interaction_data (schemas.AIInteractionLogCreate): بيانات تفاعل الذكاء الاصطناعي
        background_tasks (Optional[BackgroundTasks]): مهام خلفية

    Returns:
        models.ActivityLog: سجل النشاط
    """
    try:
        # إنشاء سجل النشاط
        log_data = schemas.ActivityLogCreate(
            log_type="ai",
            module_id=interaction_data.module_id,
            action_id=interaction_data.action_id,
            user_id=interaction_data.user_id,
            description=interaction_data.description,
            log_data={
                "agent_id": interaction_data.agent_id,
                "agent_type": interaction_data.agent_type,
                "interaction_type": interaction_data.interaction_type,
                "query": interaction_data.query,
                "response": interaction_data.response,
                "tokens_used": interaction_data.tokens_used,
                "processing_time": interaction_data.processing_time,
                "confidence_score": interaction_data.confidence_score
            },
            status=interaction_data.status
        )

        return await create_activity_log(db, log_data, background_tasks)

    except Exception as e:
        logger.error("خطأ في تسجيل تفاعل الذكاء الاصطناعي: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تسجيل تفاعل الذكاء الاصطناعي: {str(e)}") from e


# Aliases for backward compatibility
async def create_user_action_log(
    db: Session,
    action_data: Union[schemas.UserActionLogCreate, schemas.ActivityLogCreate],
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    Alias for log_user_action for backward compatibility
    """
    return await log_user_action(db, action_data, background_tasks)


async def create_ai_interaction_log(
    db: Session,
    interaction_data: schemas.AIInteractionLogCreate,
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    Alias for log_ai_interaction for backward compatibility
    """
    return await log_ai_interaction(db, interaction_data, background_tasks)


# General log_activity function for backward compatibility
async def log_activity(
    db: Session,
    log_type: str,
    module_id: str,
    action_id: str,
    description: str,
    user_id: Optional[int] = None,
    details: Optional[Dict[str, Any]] = None,
    ip_address: Optional[str] = None,
    user_agent: Optional[str] = None,
    status: str = "info",
    background_tasks: Optional[BackgroundTasks] = None
) -> models.ActivityLog:
    """
    دالة عامة لتسجيل النشاط (للتوافق مع الإصدارات السابقة)

    Args:
        db (Session): جلسة قاعدة البيانات
        log_type (str): نوع السجل
        module_id (str): معرف المديول
        action_id (str): معرف الإجراء
        description (str): الوصف
        user_id (Optional[int]): معرف المستخدم
        details (Optional[Dict[str, Any]]): تفاصيل إضافية
        ip_address (Optional[str]): عنوان IP
        user_agent (Optional[str]): وكيل المستخدم
        status (str): حالة السجل
        background_tasks (Optional[BackgroundTasks]): مهام خلفية

    Returns:
        models.ActivityLog: سجل النشاط المنشأ
    """
    try:
        # إنشاء سجل النشاط
        log_data = schemas.ActivityLogCreate(
            log_type=log_type,
            module_id=module_id,
            action_id=action_id,
            user_id=user_id,
            description=description,
            log_data=details or {},
            ip_address=ip_address,
            user_agent=user_agent,
            status=status
        )

        return await create_activity_log(db, log_data, background_tasks)

    except Exception as e:
        logger.error("خطأ في تسجيل النشاط: %s", str(e))
        raise HTTPException(status_code=500, detail=f"خطأ في تسجيل النشاط: {str(e)}") from e
