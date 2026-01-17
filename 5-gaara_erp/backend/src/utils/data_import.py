#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P2.67: Data Import from CSV/Excel

Utilities for importing data from various file formats.
"""

import io
import csv
import json
import logging
from typing import List, Dict, Any, Optional, Callable, Type
from dataclasses import dataclass, field
from datetime import datetime

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result of an import operation."""

    success: bool
    total_rows: int
    imported: int
    skipped: int
    errors: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "total_rows": self.total_rows,
            "imported": self.imported,
            "skipped": self.skipped,
            "error_count": len(self.errors),
            "errors": self.errors[:50],  # Limit errors in response
            "warnings": self.warnings,
        }


@dataclass
class ColumnMapping:
    """Mapping between file column and database field."""

    source_column: str
    target_field: str
    required: bool = False
    default_value: Any = None
    transform: Optional[Callable[[Any], Any]] = None
    validator: Optional[Callable[[Any], bool]] = None


class DataImporter:
    """
    P2.67: Generic data importer for CSV and Excel files.

    Supports:
    - CSV files
    - Excel files (xlsx, xls)
    - Column mapping
    - Data transformation
    - Validation
    - Error handling
    """

    def __init__(
        self,
        model: Type,
        column_mappings: List[ColumnMapping],
        unique_fields: List[str] = None,
        update_existing: bool = False,
    ):
        """
        Initialize the importer.

        Args:
            model: SQLAlchemy model class
            column_mappings: List of column mappings
            unique_fields: Fields to check for duplicates
            update_existing: Update if duplicate found
        """
        self.model = model
        self.column_mappings = column_mappings
        self.unique_fields = unique_fields or []
        self.update_existing = update_existing
        self._mapping_dict = {m.source_column.lower(): m for m in column_mappings}

    def import_csv(
        self, file_content: bytes, encoding: str = "utf-8", delimiter: str = ","
    ) -> ImportResult:
        """Import data from CSV file."""
        try:
            content = file_content.decode(encoding)
            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            rows = list(reader)
            return self._import_rows(rows)
        except Exception as e:
            logger.error(f"CSV import error: {e}")
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[{"row": 0, "error": str(e)}],
            )

    def import_excel(self, file_content: bytes, sheet_name: str = None) -> ImportResult:
        """Import data from Excel file."""
        try:
            import openpyxl
        except ImportError:
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[{"row": 0, "error": "openpyxl not installed"}],
            )

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            sheet = workbook[sheet_name] if sheet_name else workbook.active

            # Get headers from first row
            headers = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]

            # Convert rows to dictionaries
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {
                    headers[i]: row[i] for i in range(len(headers)) if i < len(row)
                }
                rows.append(row_dict)

            return self._import_rows(rows)
        except Exception as e:
            logger.error(f"Excel import error: {e}")
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[{"row": 0, "error": str(e)}],
            )

    def import_json(self, file_content: bytes) -> ImportResult:
        """Import data from JSON file."""
        try:
            data = json.loads(file_content.decode("utf-8"))

            if isinstance(data, dict):
                # If it's a dict, look for a data key or treat as single record
                if "data" in data:
                    rows = data["data"]
                else:
                    rows = [data]
            else:
                rows = data

            return self._import_rows(rows)
        except Exception as e:
            logger.error(f"JSON import error: {e}")
            return ImportResult(
                success=False,
                total_rows=0,
                imported=0,
                skipped=0,
                errors=[{"row": 0, "error": str(e)}],
            )

    def _import_rows(self, rows: List[Dict[str, Any]]) -> ImportResult:
        """Import a list of row dictionaries."""
        from src.database import db

        result = ImportResult(
            success=True,
            total_rows=len(rows),
            imported=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for row_num, row in enumerate(rows, start=2):  # Start at 2 (after header)
            try:
                # Transform row to model data
                model_data = self._transform_row(row, row_num, result)

                if model_data is None:
                    result.skipped += 1
                    continue

                # Check for duplicates
                existing = self._find_existing(model_data)

                if existing:
                    if self.update_existing:
                        # Update existing record
                        for key, value in model_data.items():
                            setattr(existing, key, value)
                        result.imported += 1
                        result.warnings.append(
                            f"Row {row_num}: Updated existing record"
                        )
                    else:
                        result.skipped += 1
                        result.warnings.append(
                            f"Row {row_num}: Duplicate found, skipped"
                        )
                else:
                    # Create new record
                    instance = self.model(**model_data)
                    db.session.add(instance)
                    result.imported += 1

            except Exception as e:
                result.errors.append(
                    {
                        "row": row_num,
                        "error": str(e),
                        "data": {k: str(v)[:50] for k, v in row.items()},
                    }
                )
                result.skipped += 1

        # Commit all changes
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            result.success = False
            result.errors.append({"row": 0, "error": f"Database commit failed: {e}"})
            result.imported = 0

        result.success = (
            result.imported > 0 and len(result.errors) < result.total_rows / 2
        )

        logger.info(
            f"P2.67: Import completed - {result.imported}/{result.total_rows} rows imported"
        )
        return result

    def _transform_row(
        self, row: Dict[str, Any], row_num: int, result: ImportResult
    ) -> Optional[Dict[str, Any]]:
        """Transform a row to model data using column mappings."""
        model_data = {}

        # Normalize row keys to lowercase
        row_lower = {k.lower().strip(): v for k, v in row.items() if k}

        for mapping in self.column_mappings:
            source_key = mapping.source_column.lower()

            # Get value from row
            if source_key in row_lower:
                value = row_lower[source_key]
            else:
                if mapping.required:
                    result.errors.append(
                        {
                            "row": row_num,
                            "error": f'Required column "{mapping.source_column}" not found',
                        }
                    )
                    return None
                value = mapping.default_value

            # Handle empty values
            if value is None or (isinstance(value, str) and value.strip() == ""):
                if mapping.required:
                    result.errors.append(
                        {
                            "row": row_num,
                            "error": f'Required field "{mapping.source_column}" is empty',
                        }
                    )
                    return None
                value = mapping.default_value

            # Transform value
            if value is not None and mapping.transform:
                try:
                    value = mapping.transform(value)
                except Exception as e:
                    result.errors.append(
                        {
                            "row": row_num,
                            "error": f'Transform error for "{mapping.source_column}": {e}',
                        }
                    )
                    return None

            # Validate value
            if value is not None and mapping.validator:
                if not mapping.validator(value):
                    result.errors.append(
                        {
                            "row": row_num,
                            "error": f'Validation failed for "{mapping.source_column}"',
                        }
                    )
                    return None

            model_data[mapping.target_field] = value

        return model_data

    def _find_existing(self, model_data: Dict[str, Any]) -> Optional[Any]:
        """Find existing record based on unique fields."""
        if not self.unique_fields:
            return None

        filters = {
            field: model_data.get(field)
            for field in self.unique_fields
            if model_data.get(field) is not None
        }

        if not filters:
            return None

        return self.model.query.filter_by(**filters).first()


# =============================================================================
# Transform Functions
# =============================================================================


def to_int(value: Any) -> int:
    """Convert value to integer."""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(str(value).strip().replace(",", ""))


def to_float(value: Any) -> float:
    """Convert value to float."""
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip().replace(",", ""))


def to_bool(value: Any) -> bool:
    """Convert value to boolean."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).lower().strip() in ("true", "1", "yes", "y", "نعم")


def to_date(value: Any) -> datetime:
    """Convert value to datetime."""
    if isinstance(value, datetime):
        return value

    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y"]
    value_str = str(value).strip()

    for fmt in formats:
        try:
            return datetime.strptime(value_str, fmt)
        except ValueError:
            continue

    raise ValueError(f"Cannot parse date: {value}")


def to_string(value: Any) -> str:
    """Convert value to trimmed string."""
    if value is None:
        return ""
    return str(value).strip()


# =============================================================================
# Predefined Importers
# =============================================================================


def get_products_importer():
    """Get importer for products."""
    from src.models.product import Product

    return DataImporter(
        model=Product,
        column_mappings=[
            ColumnMapping("name", "name", required=True, transform=to_string),
            ColumnMapping("sku", "sku", transform=to_string),
            ColumnMapping("barcode", "barcode", transform=to_string),
            ColumnMapping("description", "description", transform=to_string),
            ColumnMapping("price", "price", required=True, transform=to_float),
            ColumnMapping("cost", "cost", transform=to_float, default_value=0),
            ColumnMapping("quantity", "quantity", transform=to_int, default_value=0),
            ColumnMapping(
                "min_stock", "min_stock_level", transform=to_int, default_value=10
            ),
            ColumnMapping("category_id", "category_id", transform=to_int),
        ],
        unique_fields=["sku", "barcode"],
        update_existing=False,
    )


def get_customers_importer():
    """Get importer for customers."""
    from src.models.partners import Customer

    return DataImporter(
        model=Customer,
        column_mappings=[
            ColumnMapping("name", "name", required=True, transform=to_string),
            ColumnMapping("email", "email", transform=to_string),
            ColumnMapping("phone", "phone", transform=to_string),
            ColumnMapping("address", "address", transform=to_string),
            ColumnMapping("balance", "balance", transform=to_float, default_value=0),
            ColumnMapping(
                "credit_limit", "credit_limit", transform=to_float, default_value=0
            ),
        ],
        unique_fields=["email", "phone"],
        update_existing=False,
    )


def get_suppliers_importer():
    """Get importer for suppliers."""
    from src.models.partners import Supplier

    return DataImporter(
        model=Supplier,
        column_mappings=[
            ColumnMapping("name", "name", required=True, transform=to_string),
            ColumnMapping("email", "email", transform=to_string),
            ColumnMapping("phone", "phone", transform=to_string),
            ColumnMapping("address", "address", transform=to_string),
            ColumnMapping("contact_person", "contact_person", transform=to_string),
            ColumnMapping("balance", "balance", transform=to_float, default_value=0),
        ],
        unique_fields=["email"],
        update_existing=False,
    )


__all__ = [
    "DataImporter",
    "ColumnMapping",
    "ImportResult",
    "to_int",
    "to_float",
    "to_bool",
    "to_date",
    "to_string",
    "get_products_importer",
    "get_customers_importer",
    "get_suppliers_importer",
]
