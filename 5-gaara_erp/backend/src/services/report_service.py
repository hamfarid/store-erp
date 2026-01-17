#!/usr/bin/env python3
# type: ignore
# pylint: disable=all
# flake8: noqa
"""
خدمة التقارير المتقدمة
/home/ubuntu/upload/store_v1.1/complete_inventory_system/backend/src/services/report_service.py

خدمة شاملة لإنشاء وإدارة التقارير المختلفة في النظام
"""

import os
import io
import json
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
from enum import Enum

# تقارير PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# تقارير Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# قاعدة البيانات
from sqlalchemy import text
from flask import current_app

# النماذج
from src.models.inventory import Product, StockMovement, Warehouse
from src.models.customer import Customer
from src.models.supplier import Supplier
from src.models.user import User
from src.models.accounting_system import Invoice, InvoiceItem


class ReportFormat(Enum):
    """تنسيقات التقارير المدعومة"""

    PDF = "pd"
    EXCEL = "xlsx"
    CSV = "csv"
    JSON = "json"
    HTML = "html"


class ReportType(Enum):
    """أنواع التقارير المختلفة"""

    INVENTORY = "inventory"
    SALES = "sales"
    PURCHASES = "purchases"
    FINANCIAL = "financial"
    CUSTOMERS = "customers"
    SUPPLIERS = "suppliers"
    USERS = "users"
    CUSTOM = "custom"


@dataclass
class ReportConfig:
    """إعدادات التقرير"""

    title: str
    subtitle: Optional[str] = None
    date_from: Optional[datetime] = None
    date_to: Optional[datetime] = None
    filters: Optional[Dict[str, Any]] = None
    grouping: Optional[List[str]] = None
    sorting: Optional[List[str]] = None
    include_charts: bool = False
    include_summary: bool = True
    page_orientation: str = "portrait"  # portrait, landscape
    font_size: int = 10
    show_logo: bool = True
    watermark: Optional[str] = None


class ReportService:
    """خدمة التقارير الشاملة"""

    def __init__(self, db_session):
        self.db = db_session
        self.reports_dir = os.path.join(current_app.root_path, "static", "reports")
        self.templates_dir = os.path.join(current_app.root_path, "templates", "reports")

        # إنشاء المجلدات إذا لم تكن موجودة
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.templates_dir, exist_ok=True)

        # تسجيل الخطوط العربية
        self._register_arabic_fonts()

    def _register_arabic_fonts(self):
        """تسجيل الخطوط العربية لـ PDF"""
        try:
            # يمكن إضافة خطوط عربية هنا
            pass
        except Exception as e:
            print(f"Warning: Could not register Arabic fonts: {e}")

    # ==================== تقارير المخزون ====================

    def generate_inventory_report(
        self, config: ReportConfig, format: ReportFormat = ReportFormat.PDF
    ) -> str:
        """إنشاء تقرير المخزون"""

        # استعلام البيانات
        query = """
        SELECT
            p.id,
            p.name as product_name,
            p.sku,
            p.barcode,
            p.category,
            p.unit_of_measure,
            p.cost_price,
            p.selling_price,
            p.current_stock,
            p.minimum_stock,
            p.maximum_stock,
            w.name as warehouse_name,
            CASE
                WHEN p.current_stock <= p.minimum_stock THEN 'منخفض'
                WHEN p.current_stock >= p.maximum_stock THEN 'مرتفع'
                ELSE 'طبيعي'
            END as stock_status,
            (p.current_stock * p.cost_price) as total_value
        FROM products p
        LEFT JOIN warehouses w ON p.warehouse_id = w.id
        WHERE 1=1
        """

        # إضافة الفلاتر
        params = {}
        if config.filters:
            if "warehouse_id" in config.filters:
                query += " AND p.warehouse_id = :warehouse_id"
                params["warehouse_id"] = config.filters["warehouse_id"]

            if "category" in config.filters:
                query += " AND p.category = :category"
                params["category"] = config.filters["category"]

            if "stock_status" in config.filters:
                if config.filters["stock_status"] == "low":
                    query += " AND p.current_stock <= p.minimum_stock"
                elif config.filters["stock_status"] == "high":
                    query += " AND p.current_stock >= p.maximum_stock"

        # ترتيب النتائج
        if config.sorting:
            query += f" ORDER BY {', '.join(config.sorting)}"
        else:
            query += " ORDER BY p.name"

        # تنفيذ الاستعلام
        result = self.db.execute(text(query), params)
        data = result.fetchall()

        # تحويل إلى DataFrame
        df = pd.DataFrame(
            data,
            columns=[
                "ID",
                "اسم المنتج",
                "رمز المنتج",
                "الباركود",
                "الفئة",
                "وحدة القياس",
                "سعر التكلفة",
                "سعر البيع",
                "المخزون الحالي",
                "الحد الأدنى",
                "الحد الأقصى",
                "المخزن",
                "حالة المخزون",
                "القيمة الإجمالية",
            ],
        )

        # إنشاء التقرير حسب التنسيق
        if format == ReportFormat.PDF:
            return self._create_pdf_report(df, config, "تقرير المخزون")
        elif format == ReportFormat.EXCEL:
            return self._create_excel_report(df, config, "تقرير المخزون")
        elif format == ReportFormat.CSV:
            return self._create_csv_report(df, config, "تقرير المخزون")
        else:
            raise ValueError(f"Unsupported format: {format}")

    def generate_stock_movement_report(
        self, config: ReportConfig, format: ReportFormat = ReportFormat.PDF
    ) -> str:
        """تقرير حركة المخزون"""

        query = """
        SELECT
            sm.id,
            sm.movement_date,
            sm.movement_type,
            sm.quantity,
            sm.unit_cost,
            sm.total_cost,
            sm.reference_number,
            sm.notes,
            p.name as product_name,
            p.sku,
            w.name as warehouse_name,
            u.full_name as user_name
        FROM stock_movements sm
        JOIN products p ON sm.product_id = p.id
        JOIN warehouses w ON sm.warehouse_id = w.id
        LEFT JOIN users u ON sm.user_id = u.id
        WHERE 1=1
        """

        params = {}

        # فلتر التاريخ
        if config.date_from:
            query += " AND sm.movement_date >= :date_from"
            params["date_from"] = config.date_from

        if config.date_to:
            query += " AND sm.movement_date <= :date_to"
            params["date_to"] = config.date_to

        # فلاتر إضافية
        if config.filters:
            if "product_id" in config.filters:
                query += " AND sm.product_id = :product_id"
                params["product_id"] = config.filters["product_id"]

            if "warehouse_id" in config.filters:
                query += " AND sm.warehouse_id = :warehouse_id"
                params["warehouse_id"] = config.filters["warehouse_id"]

            if "movement_type" in config.filters:
                query += " AND sm.movement_type = :movement_type"
                params["movement_type"] = config.filters["movement_type"]

        query += " ORDER BY sm.movement_date DESC, sm.id DESC"

        result = self.db.execute(text(query), params)
        data = result.fetchall()

        df = pd.DataFrame(
            data,
            columns=[
                "ID",
                "تاريخ الحركة",
                "نوع الحركة",
                "الكمية",
                "سعر الوحدة",
                "التكلفة الإجمالية",
                "رقم المرجع",
                "ملاحظات",
                "اسم المنتج",
                "رمز المنتج",
                "المخزن",
                "المستخدم",
            ],
        )

        if format == ReportFormat.PDF:
            return self._create_pdf_report(df, config, "تقرير حركة المخزون")
        elif format == ReportFormat.EXCEL:
            return self._create_excel_report(df, config, "تقرير حركة المخزون")
        else:
            return self._create_csv_report(df, config, "تقرير حركة المخزون")

    # ==================== تقارير المبيعات ====================

    def generate_sales_report(
        self, config: ReportConfig, format: ReportFormat = ReportFormat.PDF
    ) -> str:
        """تقرير المبيعات"""

        query = """
        SELECT
            i.id,
            i.invoice_number,
            i.invoice_date,
            i.due_date,
            i.status,
            i.subtotal,
            i.tax_amount,
            i.discount_amount,
            i.total_amount,
            i.paid_amount,
            (i.total_amount - i.paid_amount) as remaining_amount,
            c.name as customer_name,
            c.phone as customer_phone,
            u.full_name as sales_person,
            COUNT(ii.id) as items_count
        FROM invoices i
        LEFT JOIN customers c ON i.customer_id = c.id
        LEFT JOIN users u ON i.sales_person_id = u.id
        LEFT JOIN invoice_items ii ON i.id = ii.invoice_id
        WHERE i.invoice_type = 'sale'
        """

        params = {}

        # فلتر التاريخ
        if config.date_from:
            query += " AND i.invoice_date >= :date_from"
            params["date_from"] = config.date_from

        if config.date_to:
            query += " AND i.invoice_date <= :date_to"
            params["date_to"] = config.date_to

        # فلاتر إضافية
        if config.filters:
            if "customer_id" in config.filters:
                query += " AND i.customer_id = :customer_id"
                params["customer_id"] = config.filters["customer_id"]

            if "sales_person_id" in config.filters:
                query += " AND i.sales_person_id = :sales_person_id"
                params["sales_person_id"] = config.filters["sales_person_id"]

            if "status" in config.filters:
                query += " AND i.status = :status"
                params["status"] = config.filters["status"]

        query += " GROUP BY i.id ORDER BY i.invoice_date DESC"

        result = self.db.execute(text(query), params)
        data = result.fetchall()

        df = pd.DataFrame(
            data,
            columns=[
                "ID",
                "رقم الفاتورة",
                "تاريخ الفاتورة",
                "تاريخ الاستحقاق",
                "الحالة",
                "المبلغ الفرعي",
                "الضريبة",
                "الخصم",
                "المبلغ الإجمالي",
                "المبلغ المدفوع",
                "المبلغ المتبقي",
                "اسم العميل",
                "هاتف العميل",
                "مندوب المبيعات",
                "عدد الأصناف",
            ],
        )

        if format == ReportFormat.PDF:
            return self._create_pdf_report(df, config, "تقرير المبيعات")
        elif format == ReportFormat.EXCEL:
            return self._create_excel_report(df, config, "تقرير المبيعات")
        else:
            return self._create_csv_report(df, config, "تقرير المبيعات")

    def generate_top_selling_products_report(
        self, config: ReportConfig, format: ReportFormat = ReportFormat.PDF
    ) -> str:
        """تقرير أكثر المنتجات مبيعاً"""

        query = """
        SELECT
            p.id,
            p.name as product_name,
            p.sku,
            p.category,
            SUM(ii.quantity) as total_quantity_sold,
            SUM(ii.total_price) as total_sales_amount,
            AVG(ii.unit_price) as average_selling_price,
            COUNT(DISTINCT ii.invoice_id) as number_of_invoices,
            COUNT(DISTINCT i.customer_id) as number_of_customers
        FROM products p
        JOIN invoice_items ii ON p.id = ii.product_id
        JOIN invoices i ON ii.invoice_id = i.id
        WHERE i.invoice_type = 'sale' AND i.status != 'cancelled'
        """

        params = {}

        # فلتر التاريخ
        if config.date_from:
            query += " AND i.invoice_date >= :date_from"
            params["date_from"] = config.date_from

        if config.date_to:
            query += " AND i.invoice_date <= :date_to"
            params["date_to"] = config.date_to

        query += """
        GROUP BY p.id, p.name, p.sku, p.category
        ORDER BY total_quantity_sold DESC
        LIMIT 50
        """

        result = self.db.execute(text(query), params)
        data = result.fetchall()

        df = pd.DataFrame(
            data,
            columns=[
                "ID",
                "اسم المنتج",
                "رمز المنتج",
                "الفئة",
                "إجمالي الكمية المباعة",
                "إجمالي قيمة المبيعات",
                "متوسط سعر البيع",
                "عدد الفواتير",
                "عدد العملاء",
            ],
        )

        if format == ReportFormat.PDF:
            return self._create_pdf_report(df, config, "تقرير أكثر المنتجات مبيعاً")
        elif format == ReportFormat.EXCEL:
            return self._create_excel_report(df, config, "تقرير أكثر المنتجات مبيعاً")
        else:
            return self._create_csv_report(df, config, "تقرير أكثر المنتجات مبيعاً")

    # ==================== تقارير العملاء ====================

    def generate_customer_statement_report(
        self,
        customer_id: int,
        config: ReportConfig,
        format: ReportFormat = ReportFormat.PDF,
    ) -> str:
        """كشف حساب عميل"""

        # بيانات العميل
        customer_query = """
        SELECT
            c.id,
            c.name,
            c.email,
            c.phone,
            c.address,
            c.tax_number,
            c.credit_limit,
            c.payment_terms,
            COALESCE(SUM(CASE WHEN i.status != 'cancelled' THEN i.total_amount - i.paid_amount ELSE 0 END),
                0) as outstanding_balance
        FROM customers c
        LEFT JOIN invoices i ON c.id = i.customer_id AND i.invoice_type = 'sale'
        WHERE c.id = :customer_id
        GROUP BY c.id
        """

        customer_result = self.db.execute(
            text(customer_query), {"customer_id": customer_id}
        )
        customer_data = customer_result.fetchone()

        if not customer_data:
            raise ValueError(f"Customer with ID {customer_id} not found")

        # حركات العميل
        transactions_query = """
        SELECT
            i.invoice_date as transaction_date,
            i.invoice_number as reference,
            'فاتورة مبيعات' as transaction_type,
            i.total_amount as debit_amount,
            0 as credit_amount,
            i.paid_amount,
            (i.total_amount - i.paid_amount) as balance,
            i.status,
            i.notes
        FROM invoices i
        WHERE i.customer_id = :customer_id AND i.invoice_type = 'sale'

        UNION ALL

        SELECT
            p.payment_date as transaction_date,
            p.reference_number as reference,
            'دفعة' as transaction_type,
            0 as debit_amount,
            p.amount as credit_amount,
            p.amount as paid_amount,
            0 as balance,
            'مدفوع' as status,
            p.notes
        FROM payments p
        WHERE p.customer_id = :customer_id

        ORDER BY transaction_date DESC
        """

        params = {"customer_id": customer_id}

        # فلتر التاريخ
        if config.date_from:
            transactions_query = transactions_query.replace(
                "ORDER BY transaction_date DESC",
                "AND transaction_date >= :date_from ORDER BY transaction_date DESC",
            )
            params["date_from"] = config.date_from

        if config.date_to:
            transactions_query = transactions_query.replace(
                "ORDER BY transaction_date DESC",
                "AND transaction_date <= :date_to ORDER BY transaction_date DESC",
            )
            params["date_to"] = config.date_to

        result = self.db.execute(text(transactions_query), params)
        transactions_data = result.fetchall()

        # إنشاء DataFrame للحركات
        df = pd.DataFrame(
            transactions_data,
            columns=[
                "تاريخ الحركة",
                "المرجع",
                "نوع الحركة",
                "مدين",
                "دائن",
                "المبلغ المدفوع",
                "الرصيد",
                "الحالة",
                "ملاحظات",
            ],
        )

        # إضافة معلومات العميل للتقرير
        customer_info = {
            "اسم العميل": customer_data.name,
            "البريد الإلكتروني": customer_data.email or "غير محدد",
            "الهاتف": customer_data.phone or "غير محدد",
            "العنوان": customer_data.address or "غير محدد",
            "الرقم الضريبي": customer_data.tax_number or "غير محدد",
            "حد الائتمان": (
                f"{customer_data.credit_limit:,.2f}"
                if customer_data.credit_limit
                else "غير محدد"
            ),
            "شروط الدفع": customer_data.payment_terms or "غير محدد",
            "الرصيد المستحق": f"{customer_data.outstanding_balance:,.2f}",
        }

        if format == ReportFormat.PDF:
            return self._create_customer_statement_pdf(df, customer_info, config)
        elif format == ReportFormat.EXCEL:
            return self._create_customer_statement_excel(df, customer_info, config)
        else:
            return self._create_csv_report(
                df, config, f"كشف حساب العميل - {customer_data.name}"
            )

    # ==================== إنشاء ملفات التقارير ====================

    def _create_pdf_report(
        self, df: pd.DataFrame, config: ReportConfig, title: str
    ) -> str:
        """إنشاء تقرير PDF"""

        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pd"
        filepath = os.path.join(self.reports_dir, filename)

        # إنشاء المستند
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4 if config.page_orientation == "portrait" else letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        # الأنماط
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # وسط
        )

        # محتوى التقرير
        story = []

        # العنوان
        story.append(Paragraph(title, title_style))

        # العنوان الفرعي
        if config.subtitle:
            story.append(Paragraph(config.subtitle, styles["Heading2"]))

        # معلومات التقرير
        info_data = [
            ["تاريخ الإنشاء:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        if config.date_from:
            info_data.append(["من تاريخ:", config.date_from.strftime("%Y-%m-%d")])

        if config.date_to:
            info_data.append(["إلى تاريخ:", config.date_to.strftime("%Y-%m-%d")])

        info_table = Table(info_data, colWidths=[2 * inch, 3 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ]
            )
        )

        story.append(info_table)
        story.append(Spacer(1, 20))

        # جدول البيانات
        if not df.empty:
            # تحويل DataFrame إلى قائمة
            data = [df.columns.tolist()] + df.values.tolist()

            # إنشاء الجدول
            table = Table(data)
            table.setStyle(
                TableStyle(
                    [
                        # رأس الجدول
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        # محتوى الجدول
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        # تلوين الصفوف بالتناوب
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.beige, colors.white],
                        ),
                    ]
                )
            )

            story.append(table)
        else:
            story.append(Paragraph("لا توجد بيانات لعرضها", styles["Normal"]))

        # إضافة ملخص إذا كان مطلوباً
        if config.include_summary and not df.empty:
            story.append(Spacer(1, 20))
            story.append(Paragraph("ملخص التقرير", styles["Heading2"]))

            summary_data = [
                ["إجمالي السجلات:", str(len(df))],
            ]

            # إضافة إحصائيات رقمية إذا وجدت
            numeric_columns = df.select_dtypes(include=["number"]).columns
            for col in numeric_columns:
                if not df[col].empty:
                    summary_data.append([f"مجموع {col}:", f"{df[col].sum():,.2f}"])
                    summary_data.append([f"متوسط {col}:", f"{df[col].mean():,.2f}"])

            summary_table = Table(summary_data, colWidths=[2 * inch, 2 * inch])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(summary_table)

        # بناء المستند
        doc.build(story)

        return filepath

    def _create_excel_report(
        self, df: pd.DataFrame, config: ReportConfig, title: str
    ) -> str:
        """إنشاء تقرير Excel"""

        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)

        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "التقرير"

        # تنسيق العنوان
        title_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        # العنوان
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.merge_cells("A1:" + chr(65 + len(df.columns) - 1) + "1")

        # معلومات التقرير
        row = 3
        ws[f"A{row}"] = f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1

        if config.date_from:
            ws[f"A{row}"] = f"من تاريخ: {config.date_from.strftime('%Y-%m-%d')}"
            row += 1

        if config.date_to:
            ws[f"A{row}"] = f"إلى تاريخ: {config.date_to.strftime('%Y-%m-%d')}"
            row += 1

        row += 2  # مسافة فارغة

        # إضافة البيانات
        if not df.empty:
            # رؤوس الأعمدة
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # بيانات الجدول
            for r_idx, row_data in enumerate(
                dataframe_to_rows(df, index=False, header=False), row + 1
            ):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

            # تنسيق الجدول
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in ws[
                f"A{row}:" + chr(65 + len(df.columns) - 1) + str(row + len(df))
            ]:
                for cell in row:
                    cell.border = thin_border

            # ضبط عرض الأعمدة
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except BaseException:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # إضافة الرسوم البيانية إذا كان مطلوباً
            if config.include_charts:
                self._add_excel_charts(ws, df, row + len(df) + 3)

        # حفظ الملف
        wb.save(filepath)

        return filepath

    def _create_csv_report(
        self, df: pd.DataFrame, config: ReportConfig, title: str
    ) -> str:
        """إنشاء تقرير CSV"""

        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.reports_dir, filename)

        # حفظ DataFrame كـ CSV
        df.to_csv(filepath, index=False, encoding="utf-8-sig")

        return filepath

    def _create_customer_statement_pdf(
        self, df: pd.DataFrame, customer_info: Dict, config: ReportConfig
    ) -> str:
        """إنشاء كشف حساب عميل PDF"""

        filename = f"كشف_حساب_عميل_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pd"
        filepath = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # العنوان
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
            alignment=1,
        )
        story.append(Paragraph("كشف حساب عميل", title_style))

        # معلومات العميل
        customer_data = []
        for key, value in customer_info.items():
            customer_data.append([key + ":", value])

        customer_table = Table(customer_data, colWidths=[2 * inch, 4 * inch])
        customer_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(customer_table)
        story.append(Spacer(1, 30))

        # جدول الحركات
        if not df.empty:
            data = [df.columns.tolist()] + df.values.tolist()

            movements_table = Table(data)
            movements_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.beige, colors.white],
                        ),
                    ]
                )
            )

            story.append(movements_table)

        doc.build(story)
        return filepath

    def _create_customer_statement_excel(
        self, df: pd.DataFrame, customer_info: Dict, config: ReportConfig
    ) -> str:
        """إنشاء كشف حساب عميل Excel"""

        filename = f"كشف_حساب_عميل_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "كشف حساب العميل"

        # العنوان
        ws["A1"] = "كشف حساب عميل"
        ws["A1"].font = Font(size=18, bold=True)
        ws.merge_cells("A1:F1")

        # معلومات العميل
        row = 3
        for key, value in customer_info.items():
            ws[f"A{row}"] = key + ":"
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        row += 2

        # جدول الحركات
        if not df.empty:
            # رؤوس الأعمدة
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )

            # البيانات
            for r_idx, row_data in enumerate(
                dataframe_to_rows(df, index=False, header=False), row + 1
            ):
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(filepath)
        return filepath

    def _add_excel_charts(self, ws, df: pd.DataFrame, start_row: int):
        """إضافة الرسوم البيانية لـ Excel"""

        # البحث عن الأعمدة الرقمية
        numeric_columns = df.select_dtypes(include=["number"]).columns

        if len(numeric_columns) > 0:
            # رسم بياني عمودي
            chart = BarChart()
            chart.title = "الرسم البياني"
            chart.x_axis.title = "الفئات"
            chart.y_axis.title = "القيم"

            # إضافة البيانات
            data = Reference(
                ws,
                min_col=2,
                min_row=7,
                max_row=7 + len(df),
                max_col=2 + len(numeric_columns) - 1,
            )
            cats = Reference(ws, min_col=1, min_row=8, max_row=7 + len(df))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # إضافة الرسم للورقة
            ws.add_chart(chart, f"A{start_row}")

    # ==================== وظائف مساعدة ====================

    def get_available_reports(self) -> List[Dict[str, Any]]:
        """الحصول على قائمة التقارير المتاحة"""

        reports = [
            {
                "id": "inventory_report",
                "name": "تقرير المخزون",
                "description": "تقرير شامل عن حالة المخزون الحالية",
                "category": "inventory",
                "filters": ["warehouse_id", "category", "stock_status"],
                "formats": ["pd", "excel", "csv"],
            },
            {
                "id": "stock_movement_report",
                "name": "تقرير حركة المخزون",
                "description": "تقرير تفصيلي عن حركات المخزون",
                "category": "inventory",
                "filters": [
                    "date_from",
                    "date_to",
                    "product_id",
                    "warehouse_id",
                    "movement_type",
                ],
                "formats": ["pd", "excel", "csv"],
            },
            {
                "id": "sales_report",
                "name": "تقرير المبيعات",
                "description": "تقرير شامل عن المبيعات والفواتير",
                "category": "sales",
                "filters": [
                    "date_from",
                    "date_to",
                    "customer_id",
                    "sales_person_id",
                    "status",
                ],
                "formats": ["pd", "excel", "csv"],
            },
            {
                "id": "top_selling_products_report",
                "name": "تقرير أكثر المنتجات مبيعاً",
                "description": "تقرير عن أكثر المنتجات مبيعاً خلال فترة محددة",
                "category": "sales",
                "filters": ["date_from", "date_to"],
                "formats": ["pd", "excel", "csv"],
            },
            {
                "id": "customer_statement_report",
                "name": "كشف حساب عميل",
                "description": "كشف حساب تفصيلي لعميل محدد",
                "category": "customers",
                "filters": ["customer_id", "date_from", "date_to"],
                "formats": ["pd", "excel", "csv"],
            },
        ]

        return reports

    def delete_old_reports(self, days_old: int = 30):
        """حذف التقارير القديمة"""

        cutoff_date = datetime.now() - timedelta(days=days_old)

        for filename in os.listdir(self.reports_dir):
            filepath = os.path.join(self.reports_dir, filename)

            if os.path.isfile(filepath):
                file_modified = datetime.fromtimestamp(os.path.getmtime(filepath))

                if file_modified < cutoff_date:
                    try:
                        os.remove(filepath)
                        print(f"Deleted old report: {filename}")
                    except Exception as e:
                        print(f"Error deleting {filename}: {e}")

    def get_report_file_url(self, filename: str) -> str:
        """الحصول على رابط ملف التقرير"""
        return f"/static/reports/{filename}"


# ==================== Factory Functions ====================


def create_report_service(db_session):
    """إنشاء خدمة التقارير"""
    return ReportService(db_session)


def generate_report(
    report_type: str,
    config: ReportConfig,
    format: ReportFormat = ReportFormat.PDF,
    **kwargs,
):
    """دالة مساعدة لإنشاء التقارير"""

    from app import db  # استيراد قاعدة البيانات من التطبيق

    service = create_report_service(db.session)

    if report_type == "inventory":
        return service.generate_inventory_report(config, format)
    elif report_type == "stock_movement":
        return service.generate_stock_movement_report(config, format)
    elif report_type == "sales":
        return service.generate_sales_report(config, format)
    elif report_type == "top_selling_products":
        return service.generate_top_selling_products_report(config, format)
    elif report_type == "customer_statement":
        customer_id = kwargs.get("customer_id")
        if not customer_id:
            raise ValueError("customer_id is required for customer statement report")
        return service.generate_customer_statement_report(customer_id, config, format)
    else:
        raise ValueError(f"Unknown report type: {report_type}")


if __name__ == "__main__":
    # اختبار الخدمة
    print("Report Service initialized successfully!")
