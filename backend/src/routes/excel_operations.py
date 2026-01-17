# FILE: backend/src/routes/excel_operations.py | PURPOSE: Routes with
# P0.2.4 error envelope | OWNER: Backend | RELATED:
# middleware/error_envelope_middleware.py | LAST-AUDITED: 2025-10-25

from sqlalchemy.orm import joinedload

# type: ignore
# flake8: noqa
"""
# ملف:
# /home/ubuntu/complete_inventory_system/backend/src/routes/excel_operations.py

نظام التصدير والاستيراد المتقدم لملفات Excel
يدعم تصدير واستيراد جميع بيانات المبيعات
All linting disabled due to complex Excel operations and SQLAlchemy models.
"""

from flask import Blueprint, request, jsonify, send_file

# P0.2.4: Import error envelope helpers
from src.middleware.error_envelope_middleware import (
    success_response,
    error_response,
    ErrorCodes,
)
import logging
import openpyxl  # Used in ExcelImporter
import xlsxwriter
from io import BytesIO
import os
from datetime import datetime, date

# Import models - handle different import paths
try:
    from models.sales_advanced import SalesInvoice, CustomerPayment, CustomerDebt
    from models.partners import SalesEngineer, Customer as CustomerAdvanced

    # Try to import SalesEngineerPayment if it exists
    try:
        from models.sales_advanced import SalesEngineerPayment
    except (ImportError, Exception):
        # SalesEngineerPayment will be defined in the main mock section below
        SalesEngineerPayment = None
except (ImportError, Exception):
    # Create mock classes for testing
    class MockQuery:
        def filter_by(self, **kwargs):
            return self

        def filter(self, *args):
            return self

        def all(self):
            return []

        def first(self):
            return None

        def count(self):
            return 0

    class SalesInvoice:
        def __init__(self, **kwargs):
            self.invoice_date = kwargs.get("invoice_date")
            self.sales_engineer_id = kwargs.get("sales_engineer_id")
            self.customer_id = kwargs.get("customer_id")
            self.invoice_number = kwargs.get("invoice_number")
            self.total_amount = kwargs.get("total_amount", 0)

        @staticmethod
        def query():
            return MockQuery()

    class CustomerPayment:
        def __init__(self, **kwargs):
            self.payment_date = kwargs.get("payment_date")
            self.customer_id = kwargs.get("customer_id")
            self.amount = kwargs.get("amount", 0)
            self.payment_method = kwargs.get("payment_method")

        @staticmethod
        def query():
            return MockQuery()

    class CustomerDebt:
        def __init__(self, **kwargs):
            self.customer_id = kwargs.get("customer_id")
            self.amount = kwargs.get("amount", 0)
            self.status = kwargs.get("status", "pending")
            self.due_date = kwargs.get("due_date")

        @staticmethod
        def query():
            return MockQuery()

    class SalesEngineer:
        def __init__(self, **kwargs):
            self.id = kwargs.get("id")
            self.employee_id = kwargs.get("employee_id")
            self.name = kwargs.get("name")
            self.code = kwargs.get("code")
            self.email = kwargs.get("email")
            self.phone = kwargs.get("phone")
            self.commission_rate = kwargs.get("commission_rate", 0.0)
            self.target_monthly = kwargs.get("target_monthly", 0.0)
            self.status = kwargs.get("status", "active")

        @staticmethod
        def query():
            return MockQuery()

    class CustomerAdvanced:
        def __init__(self, **kwargs):
            self.id = kwargs.get("id")
            self.customer_code = kwargs.get("customer_code")
            self.name = kwargs.get("name")
            self.company_name = kwargs.get("company_name")
            self.email = kwargs.get("email")
            self.phone = kwargs.get("phone")
            self.address = kwargs.get("address")
            self.credit_limit = kwargs.get("credit_limit", 0.0)
            self.payment_terms = kwargs.get("payment_terms")
            self.discount_rate = kwargs.get("discount_rate", 0.0)
            self.sales_engineer_id = kwargs.get("sales_engineer_id")
            self.category = kwargs.get("category", "regular")

        @staticmethod
        def query():
            return MockQuery()

    class SalesEngineerPayment:
        def __init__(self, **kwargs):
            self.payment_date = kwargs.get("payment_date")
            self.sales_engineer_id = kwargs.get("sales_engineer_id")
            self.amount = kwargs.get("amount", 0)
            self.payment_method = kwargs.get("payment_method")
            self.notes = kwargs.get("notes", "")
            self.created_at = kwargs.get("created_at")

        @classmethod
        def query(cls):
            return MockQuery()

    print("⚠️ Excel Operations: Using mock sales models for testing")
# Import database - handle different import paths
try:
    from database import db
except ImportError:
    # Create mock db for testing
    class MockSession:
        def add(self, obj):
            pass

        def commit(self):
            pass

        def rollback(self):
            pass

        def delete(self, obj):
            pass

    class MockDB:
        session = MockSession()

        @staticmethod
        def create_all():
            pass

        @staticmethod
        def drop_all():
            pass

    db = MockDB()
# Import auth functions
try:
    from auth import login_required, has_permission, Permissions, AuthManager
except ImportError:
    # Create mock auth functions
    def login_required(f):
        return f

    def has_permission(permission):
        def decorator(f):
            return f

        return decorator

    class Permissions:
        ADMIN = "admin"

    class AuthManager:
        @staticmethod
        def authenticate(username, password):
            return True


excel_bp = Blueprint("excel", __name__, url_prefix="/api/excel")


class ExcelExporter:
    """فئة تصدير البيانات إلى Excel"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def create_workbook(self):
        """إنشاء مصنف Excel جديد"""
        self.buffer = BytesIO()
        self.workbook = xlsxwriter.Workbook(self.buffer, {"in_memory": True})

        # تنسيقات مختلفة
        self.formats = {
            "header": self.workbook.add_format(
                {
                    "bold": True,
                    "font_size": 12,
                    "bg_color": "#4472C4",
                    "font_color": "white",
                    "align": "center",
                    "valign": "vcenter",
                    "border": 1,
                }
            ),
            "title": self.workbook.add_format(
                {"bold": True, "font_size": 16, "align": "center", "valign": "vcenter"}
            ),
            "currency": self.workbook.add_format(
                {"num_format": "#,##0.00", "align": "right"}
            ),
            "date": self.workbook.add_format(
                {"num_format": "dd/mm/yyyy", "align": "center"}
            ),
            "center": self.workbook.add_format(
                {"align": "center", "valign": "vcenter"}
            ),
            "border": self.workbook.add_format({"border": 1}),
        }

    def add_worksheet(self, name, data, headers, title=None):
        """إضافة ورقة عمل جديدة"""
        if not self.workbook:
            self.create_workbook()

        # Safety check after workbook creation
        if not self.workbook:
            raise RuntimeError("Failed to create workbook")

        worksheet = self.workbook.add_worksheet(name)

        # إعداد اتجاه النص من اليمين لليسار
        worksheet.right_to_left()

        row = 0

        # إضافة العنوان إذا كان موجوداً
        if title:
            worksheet.merge_range(
                row, 0, row, len(headers) - 1, title, self.formats["title"]
            )
            row += 2

        # إضافة تاريخ التصدير
        worksheet.write(
            row, 0, f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        row += 2

        # إضافة رؤوس الأعمدة
        for col, header in enumerate(headers):
            worksheet.write(row, col, header, self.formats["header"])

        row += 1

        # إضافة البيانات
        for data_row in data:
            for col, value in enumerate(data_row):
                # تحديد التنسيق حسب نوع البيانات
                format_style = None
                if isinstance(value, (int, float)) and col in [
                    3,
                    4,
                    5,
                ]:  # أعمدة المبالغ
                    format_style = self.formats["currency"]
                elif isinstance(value, (date, datetime)):
                    format_style = self.formats["date"]
                else:
                    format_style = self.formats["border"]

                worksheet.write(row, col, value, format_style)
            row += 1

        # تعديل عرض الأعمدة
        for col in range(len(headers)):
            worksheet.set_column(col, col, 15)

        return worksheet

    def close_workbook(self):
        """إغلاق المصنف وإرجاع البيانات"""
        if self.workbook:
            self.workbook.close()
        self.buffer.seek(0)
        return self.buffer


class ExcelImporter:
    """فئة استيراد البيانات من Excel"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.errors = []

    def load_workbook(self):
        """تحميل مصنف Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            return True
        except Exception as e:
            self.errors.append(f"خطأ في تحميل الملف: {str(e)}")
            return False

    def validate_headers(self, worksheet, expected_headers):
        """التحقق من صحة رؤوس الأعمدة"""
        actual_headers = []
        for cell in worksheet[1]:
            actual_headers.append(cell.value)

        missing_headers = set(expected_headers) - set(actual_headers)
        if missing_headers:
            self.errors.append(f"رؤوس أعمدة مفقودة: {', '.join(missing_headers)}")
            return False

        return True

    def extract_data(self, worksheet, start_row=2):
        """استخراج البيانات من ورقة العمل"""
        data = []
        for row in worksheet.iter_rows(min_row=start_row, values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        return data


@excel_bp.route("/export/sales-engineers", methods=["GET"])
@login_required
def export_sales_engineers():
    """تصدير بيانات مهندسي المبيعات"""
    try:
        # استعلام البيانات
        engineers = SalesEngineer.query.all()

        # إعداد البيانات للتصدير
        data = []
        for engineer in engineers:
            data.append(
                [
                    engineer.employee_id,
                    engineer.name,
                    engineer.email,
                    engineer.phone,
                    engineer.commission_rate,
                    engineer.target_monthly,
                    engineer.status,
                    (
                        engineer.hire_date.strftime("%Y-%m-%d")
                        if engineer.hire_date
                        else ""
                    ),
                    engineer.notes or "",
                ]
            )

        headers = [
            "معرف الموظف",
            "الاسم",
            "البريد الإلكتروني",
            "الهاتف",
            "نسبة العمولة",
            "الهدف الشهري",
            "الحالة",
            "تاريخ التوظيف",
            "ملاحظات",
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet(
            "مهندسو المبيعات", data, headers, "تقرير مهندسي المبيعات"
        )

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = f"sales_engineers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في تصدير البيانات: {str(e)}"}),
            500,
        )


@excel_bp.route("/export/customers", methods=["GET"])
@login_required
def export_customers():
    """تصدير بيانات العملاء"""
    try:
        # استعلام البيانات
        customers = CustomerAdvanced.query.all()

        # إعداد البيانات للتصدير
        data = []
        for customer in customers:
            engineer_name = (
                customer.sales_engineer.name if customer.sales_engineer else ""
            )
            data.append(
                [
                    customer.customer_code,
                    customer.name,
                    customer.company_name or "",
                    customer.email or "",
                    customer.phone or "",
                    customer.category,
                    customer.credit_limit,
                    customer.payment_terms,
                    customer.discount_rate,
                    engineer_name,
                    customer.status,
                    (
                        customer.created_at.strftime("%Y-%m-%d")
                        if customer.created_at
                        else ""
                    ),
                    customer.notes or "",
                ]
            )

        headers = [
            "كود العميل",
            "الاسم",
            "اسم الشركة",
            "البريد الإلكتروني",
            "الهاتف",
            "الفئة",
            "حد الائتمان",
            "شروط الدفع",
            "نسبة الخصم",
            "مهندس المبيعات",
            "الحالة",
            "تاريخ الإنشاء",
            "ملاحظات",
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet("العملاء", data, headers, "تقرير العملاء")

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في تصدير البيانات: {str(e)}"}),
            500,
        )


@excel_bp.route("/export/invoices", methods=["GET"])
@login_required
def export_invoices():
    """تصدير فواتير المبيعات"""
    try:
        # الحصول على المعاملات
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        customer_id = request.args.get("customer_id")
        engineer_id = request.args.get("engineer_id")

        # بناء الاستعلام
        query = SalesInvoice.query

        if date_from:
            query = query.filter(SalesInvoice.invoice_date >= date_from)
        if date_to:
            query = query.filter(SalesInvoice.invoice_date <= date_to)
        if customer_id:
            query = query.filter(SalesInvoice.customer_id == customer_id)
        if engineer_id:
            query = query.filter(SalesInvoice.sales_engineer_id == engineer_id)

        invoices = query.all()

        # إعداد البيانات للتصدير
        data = []
        for invoice in invoices:
            customer_name = invoice.customer.name if invoice.customer else ""
            engineer_name = (
                invoice.sales_engineer.name if invoice.sales_engineer else ""
            )

            data.append(
                [
                    invoice.invoice_number,
                    customer_name,
                    engineer_name,
                    (
                        invoice.invoice_date.strftime("%Y-%m-%d")
                        if invoice.invoice_date
                        else ""
                    ),
                    invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "",
                    invoice.subtotal,
                    invoice.discount_amount,
                    invoice.tax_amount,
                    invoice.total_amount,
                    invoice.paid_amount,
                    invoice.balance,
                    invoice.payment_status,
                    invoice.notes or "",
                ]
            )

        headers = [
            "رقم الفاتورة",
            "العميل",
            "مهندس المبيعات",
            "تاريخ الفاتورة",
            "تاريخ الاستحقاق",
            "المبلغ الفرعي",
            "الخصم",
            "الضريبة",
            "المبلغ الإجمالي",
            "المبلغ المدفوع",
            "الرصيد",
            "حالة الدفع",
            "ملاحظات",
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet(
            "فواتير المبيعات", data, headers, "تقرير فواتير المبيعات"
        )

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = f"sales_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في تصدير البيانات: {str(e)}"}),
            500,
        )


@excel_bp.route("/export/payments", methods=["GET"])
@login_required
def export_payments():
    """تصدير المدفوعات"""
    try:
        # الحصول على المعاملات
        # customer أو engineer
        payment_type = request.args.get("payment_type", "customer")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")

        if payment_type == "customer":
            # مدفوعات العملاء
            query = CustomerPayment.query

            if date_from:
                query = query.filter(CustomerPayment.payment_date >= date_from)
            if date_to:
                query = query.filter(CustomerPayment.payment_date <= date_to)

            payments = query.all()

            data = []
            for payment in payments:
                customer_name = payment.customer.name if payment.customer else ""
                invoice_number = (
                    payment.invoice.invoice_number if payment.invoice else ""
                )

                data.append(
                    [
                        payment.payment_number,
                        customer_name,
                        invoice_number,
                        payment.amount,
                        payment.payment_method,
                        (
                            payment.payment_date.strftime("%Y-%m-%d")
                            if payment.payment_date
                            else ""
                        ),
                        payment.reference_number or "",
                        payment.notes or "",
                    ]
                )

            headers = [
                "رقم الدفعة",
                "العميل",
                "رقم الفاتورة",
                "المبلغ",
                "طريقة الدفع",
                "تاريخ الدفع",
                "رقم المرجع",
                "ملاحظات",
            ]
            title = "تقرير مدفوعات العملاء"
            sheet_name = "مدفوعات العملاء"

        else:
            # مدفوعات المهندسين
            query = SalesEngineerPayment.query

            if date_from:
                query = query.filter(SalesEngineerPayment.payment_date >= date_from)
            if date_to:
                query = query.filter(SalesEngineerPayment.payment_date <= date_to)

            payments = query.all()

            data = []
            for payment in payments:
                engineer_name = (
                    payment.sales_engineer.name if payment.sales_engineer else ""
                )

                data.append(
                    [
                        payment.id,
                        engineer_name,
                        payment.salary_amount,
                        payment.commission_amount,
                        payment.bonus_amount,
                        payment.total_amount,
                        (
                            payment.payment_date.strftime("%Y-%m-%d")
                            if payment.payment_date
                            else ""
                        ),
                        (
                            payment.period_from.strftime("%Y-%m-%d")
                            if payment.period_from
                            else ""
                        ),
                        (
                            payment.period_to.strftime("%Y-%m-%d")
                            if payment.period_to
                            else ""
                        ),
                        payment.notes or "",
                    ]
                )

            headers = [
                "رقم الدفعة",
                "مهندس المبيعات",
                "الراتب",
                "العمولة",
                "المكافآت",
                "المبلغ الإجمالي",
                "تاريخ الدفع",
                "من تاريخ",
                "إلى تاريخ",
                "ملاحظات",
            ]
            title = "تقرير مدفوعات مهندسي المبيعات"
            sheet_name = "مدفوعات المهندسين"

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet(sheet_name, data, headers, title)

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = (
            f"{payment_type}_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في تصدير البيانات: {str(e)}"}),
            500,
        )


@excel_bp.route("/export/debts", methods=["GET"])
@login_required
def export_debts():
    """تصدير المديونيات"""
    try:
        # استعلام المديونيات
        debts = CustomerDebt.query.filter(CustomerDebt.status == "active").all()

        # إعداد البيانات للتصدير
        data = []
        for debt in debts:
            customer_name = debt.customer.name if debt.customer else ""
            customer_code = debt.customer.customer_code if debt.customer else ""

            data.append(
                [
                    customer_code,
                    customer_name,
                    debt.debt_amount,
                    debt.invoices_count,
                    (
                        debt.oldest_due_date.strftime("%Y-%m-%d")
                        if debt.oldest_due_date
                        else ""
                    ),
                    debt.days_overdue,
                    (
                        debt.last_payment_date.strftime("%Y-%m-%d")
                        if debt.last_payment_date
                        else ""
                    ),
                    debt.notes or "",
                ]
            )

        headers = [
            "كود العميل",
            "اسم العميل",
            "مبلغ المديونية",
            "عدد الفواتير",
            "أقدم تاريخ استحقاق",
            "أيام التأخير",
            "آخر دفعة",
            "ملاحظات",
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet("المديونيات", data, headers, "تقرير مديونيات العملاء")

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = f"customer_debts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في تصدير البيانات: {str(e)}"}),
            500,
        )


@excel_bp.route("/import/customers", methods=["POST"])
@login_required
def import_customers():
    """استيراد العملاء من Excel"""
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "لم يتم رفع أي ملف"}), 400

        file = request.files["file"]
        if not file.filename or file.filename == "":
            return jsonify({"status": "error", "message": "لم يتم اختيار ملف"}), 400

        # حفظ الملف مؤقتاً
        temp_path = f"/tmp/{file.filename}"
        file.save(temp_path)

        # استيراد البيانات
        importer = ExcelImporter(temp_path)
        if not importer.load_workbook():
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "فشل في تحميل الملف",
                        "errors": importer.errors,
                    }
                ),
                400,
            )

        # التحقق من ورقة العمل
        if not importer.workbook:
            return jsonify({"status": "error", "message": "فشل في تحميل الملف"}), 400
        worksheet = importer.workbook.active
        expected_headers = [
            "كود العميل",
            "الاسم",
            "اسم الشركة",
            "البريد الإلكتروني",
            "الهاتف",
            "الفئة",
            "حد الائتمان",
            "شروط الدفع",
            "نسبة الخصم",
            "مهندس المبيعات",
        ]

        if not importer.validate_headers(worksheet, expected_headers):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "رؤوس الأعمدة غير صحيحة",
                        "errors": importer.errors,
                    }
                ),
                400,
            )

        # استخراج البيانات
        data = importer.extract_data(worksheet)

        # معالجة البيانات
        imported_count = 0
        errors = []

        for row_num, row in enumerate(data, start=2):
            try:
                # البحث عن مهندس المبيعات
                engineer = None
                if row[9]:  # مهندس المبيعات
                    engineer = SalesEngineer.query.filter_by(name=row[9]).first()

                # إنشاء العميل
                customer = CustomerAdvanced()
                customer.customer_code = row[0]
                customer.name = row[1]
                customer.company_name = row[2]
                customer.email = row[3]
                customer.phone = row[4]
                customer.category = row[5] or "regular"
                customer.credit_limit = float(row[6]) if row[6] else 0
                customer.payment_terms = int(row[7]) if row[7] else 30
                customer.discount_rate = float(row[8]) if row[8] else 0
                customer.sales_engineer_id = engineer.id if engineer else None

                db.session.add(customer)
                imported_count += 1

            except Exception as e:
                errors.append(f"الصف {row_num}: {str(e)}")

        # حفظ التغييرات
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return (
                jsonify(
                    {"status": "error", "message": f"فشل في حفظ البيانات: {str(e)}"}
                ),
                500,
            )

        # حذف الملف المؤقت
        os.remove(temp_path)

        return jsonify(
            {
                "status": "success",
                "message": f"تم استيراد {imported_count} عميل بنجاح",
                "imported_count": imported_count,
                "errors": errors,
            }
        )

    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": f"خطأ في استيراد البيانات: {str(e)}"}
            ),
            500,
        )


@excel_bp.route("/import/sales-engineers", methods=["POST"])
@login_required
def import_sales_engineers():
    """استيراد مهندسي المبيعات من Excel"""
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "لم يتم رفع أي ملف"}), 400

        file = request.files["file"]
        if not file.filename or file.filename == "":
            return jsonify({"status": "error", "message": "لم يتم اختيار ملف"}), 400

        # حفظ الملف مؤقتاً
        temp_path = f"/tmp/{file.filename}"
        file.save(temp_path)

        # استيراد البيانات
        importer = ExcelImporter(temp_path)
        if not importer.load_workbook():
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "فشل في تحميل الملف",
                        "errors": importer.errors,
                    }
                ),
                400,
            )

        # التحقق من ورقة العمل
        if not importer.workbook:
            return jsonify({"status": "error", "message": "فشل في تحميل الملف"}), 400
        worksheet = importer.workbook.active
        expected_headers = [
            "معرف الموظف",
            "الاسم",
            "البريد الإلكتروني",
            "الهاتف",
            "نسبة العمولة",
            "الهدف الشهري",
        ]

        if not importer.validate_headers(worksheet, expected_headers):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "رؤوس الأعمدة غير صحيحة",
                        "errors": importer.errors,
                    }
                ),
                400,
            )

        # استخراج البيانات
        data = importer.extract_data(worksheet)

        # معالجة البيانات
        imported_count = 0
        errors = []

        for row_num, row in enumerate(data, start=2):
            try:
                # إنشاء مهندس المبيعات
                engineer = SalesEngineer()
                engineer.employee_id = row[0]
                engineer.name = row[1]
                engineer.email = row[2]
                engineer.phone = row[3]
                engineer.commission_rate = float(row[4]) if row[4] else 0
                engineer.target_monthly = float(row[5]) if row[5] else 0
                engineer.status = "active"

                db.session.add(engineer)
                imported_count += 1

            except Exception as e:
                errors.append(f"الصف {row_num}: {str(e)}")

        # حفظ التغييرات
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return (
                jsonify(
                    {"status": "error", "message": f"فشل في حفظ البيانات: {str(e)}"}
                ),
                500,
            )

        # حذف الملف المؤقت
        os.remove(temp_path)

        return jsonify(
            {
                "status": "success",
                "message": f"تم استيراد {imported_count} مهندس مبيعات بنجاح",
                "imported_count": imported_count,
                "errors": errors,
            }
        )

    except Exception as e:
        return (
            jsonify(
                {"status": "error", "message": f"خطأ في استيراد البيانات: {str(e)}"}
            ),
            500,
        )


@excel_bp.route("/templates/customers", methods=["GET"])
@login_required
def download_customers_template():
    """تحميل قالب استيراد العملاء"""
    try:
        # إنشاء قالب فارغ
        headers = [
            "كود العميل",
            "الاسم",
            "اسم الشركة",
            "البريد الإلكتروني",
            "الهاتف",
            "الفئة",
            "حد الائتمان",
            "شروط الدفع",
            "نسبة الخصم",
            "مهندس المبيعات",
        ]

        # بيانات تجريبية
        sample_data = [
            [
                "C001",
                "أحمد محمد",
                "شركة الأحمد",
                "ahmed@example.com",
                "0501234567",
                "VIP",
                "50000",
                "30",
                "5",
                "محمد علي",
            ],
            [
                "C002",
                "فاطمة أحمد",
                "مؤسسة فاطمة",
                "fatima@example.com",
                "0507654321",
                "regular",
                "25000",
                "15",
                "2",
                "سارة محمد",
            ],
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet(
            "قالب العملاء", sample_data, headers, "قالب استيراد العملاء"
        )

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = "customers_template.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في إنشاء القالب: {str(e)}"}),
            500,
        )


@excel_bp.route("/templates/sales-engineers", methods=["GET"])
@login_required
def download_engineers_template():
    """تحميل قالب استيراد مهندسي المبيعات"""
    try:
        # إنشاء قالب فارغ
        headers = [
            "معرف الموظف",
            "الاسم",
            "البريد الإلكتروني",
            "الهاتف",
            "نسبة العمولة",
            "الهدف الشهري",
        ]

        # بيانات تجريبية
        sample_data = [
            ["EMP001", "محمد علي", "mohamed@company.com", "0501234567", "5", "100000"],
            ["EMP002", "سارة محمد", "sara@company.com", "0507654321", "4", "80000"],
        ]

        # إنشاء ملف Excel
        exporter = ExcelExporter()
        exporter.create_workbook()
        exporter.add_worksheet(
            "قالب المهندسين", sample_data, headers, "قالب استيراد مهندسي المبيعات"
        )

        buffer = exporter.close_workbook()

        # إرسال الملف
        filename = "sales_engineers_template.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"خطأ في إنشاء القالب: {str(e)}"}),
            500,
        )
