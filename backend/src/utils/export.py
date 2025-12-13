#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P2.56: Export to Excel/PDF Functionality

Provides export capabilities for reports and data:
- Excel (XLSX) export with formatting
- PDF export with templates
- CSV export
- JSON export
"""

import io
import os
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# =============================================================================
# Configuration
# =============================================================================

EXPORT_DIR = os.environ.get("EXPORT_DIR", "exports")
MAX_EXPORT_ROWS = int(os.environ.get("MAX_EXPORT_ROWS", 10000))


# =============================================================================
# Data Classes
# =============================================================================


@dataclass
class ExportColumn:
    """Definition of an export column."""

    key: str
    label: str
    width: int = 15
    format: Optional[str] = None  # 'currency', 'date', 'number', 'percentage'


@dataclass
class ExportConfig:
    """Configuration for export."""

    title: str
    columns: List[ExportColumn]
    filename: Optional[str] = None
    include_totals: bool = False
    totals_columns: Optional[List[str]] = None
    date_format: str = "%Y-%m-%d %H:%M:%S"
    currency_symbol: str = "$"
    rtl: bool = False  # Right-to-left for Arabic


# =============================================================================
# Excel Export
# =============================================================================


class ExcelExporter:
    """
    P2.56: Excel (XLSX) exporter with formatting support.
    """

    def __init__(self, config: ExportConfig):
        self.config = config
        self._workbook = None
        self._worksheet = None

    def export(self, data: List[Dict[str, Any]]) -> bytes:
        """
        Export data to Excel format.

        Args:
            data: List of dictionaries to export

        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl required for Excel export")

        wb = Workbook()
        ws = wb.active
        ws.title = self.config.title[:31]  # Excel limit

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write title
        ws.merge_cells(f"A1:{get_column_letter(len(self.config.columns))}1")
        title_cell = ws["A1"]
        title_cell.value = self.config.title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Write timestamp
        ws.merge_cells(f"A2:{get_column_letter(len(self.config.columns))}2")
        date_cell = ws["A2"]
        date_cell.value = (
            f"Generated: {datetime.now().strftime(self.config.date_format)}"
        )
        date_cell.alignment = Alignment(horizontal="center")

        # Write headers (row 4)
        header_row = 4
        for col_idx, column in enumerate(self.config.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=column.label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = column.width

        # Write data
        totals = {
            col.key: 0
            for col in self.config.columns
            if col.key in (self.config.totals_columns or [])
        }

        for row_idx, row_data in enumerate(data[:MAX_EXPORT_ROWS], header_row + 1):
            for col_idx, column in enumerate(self.config.columns, 1):
                value = row_data.get(column.key, "")

                # Format value
                if column.format == "currency" and isinstance(value, (int, float)):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = f"{self.config.currency_symbol}#,##0.00"
                    if column.key in totals:
                        totals[column.key] += value
                elif column.format == "date" and value:
                    if isinstance(value, str):
                        try:
                            value = datetime.fromisoformat(value)
                        except BaseException:
                            pass
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = "YYYY-MM-DD"
                elif column.format == "number" and isinstance(value, (int, float)):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = "#,##0"
                    if column.key in totals:
                        totals[column.key] += value
                elif column.format == "percentage" and isinstance(value, (int, float)):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value / 100)
                    cell.number_format = "0.00%"
                else:
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=str(value) if value else ""
                    )

                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if self.config.rtl else "left"
                )

        # Write totals
        if self.config.include_totals and totals:
            totals_row = header_row + len(data) + 1
            for col_idx, column in enumerate(self.config.columns, 1):
                if column.key in totals:
                    cell = ws.cell(
                        row=totals_row, column=col_idx, value=totals[column.key]
                    )
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    if column.format == "currency":
                        cell.number_format = f"{self.config.currency_symbol}#,##0.00"
                elif col_idx == 1:
                    cell = ws.cell(row=totals_row, column=col_idx, value="Total")
                    cell.font = Font(bold=True)
                    cell.border = thin_border

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()


# =============================================================================
# PDF Export
# =============================================================================


class PDFExporter:
    """
    P2.56: PDF exporter with table formatting.
    """

    def __init__(self, config: ExportConfig):
        self.config = config

    def export(self, data: List[Dict[str, Any]]) -> bytes:
        """
        Export data to PDF format.

        Args:
            data: List of dictionaries to export

        Returns:
            PDF file as bytes
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise ImportError("reportlab required for PDF export")

        output = io.BytesIO()

        # Create document
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "Title", parent=styles["Heading1"], alignment=1, spaceAfter=12
        )  # Center
        elements.append(Paragraph(self.config.title, title_style))

        # Date
        date_style = ParagraphStyle(
            "Date",
            parent=styles["Normal"],
            alignment=1,
            fontSize=10,
            textColor=colors.grey,
        )
        elements.append(
            Paragraph(
                f"Generated: {datetime.now().strftime(self.config.date_format)}",
                date_style,
            )
        )
        elements.append(Spacer(1, 20))

        # Table data
        table_data = []

        # Headers
        headers = [col.label for col in self.config.columns]
        table_data.append(headers)

        # Rows
        totals = {
            col.key: 0
            for col in self.config.columns
            if col.key in (self.config.totals_columns or [])
        }

        for row_data in data[:MAX_EXPORT_ROWS]:
            row = []
            for column in self.config.columns:
                value = row_data.get(column.key, "")

                if column.format == "currency" and isinstance(value, (int, float)):
                    formatted = f"{self.config.currency_symbol}{value:,.2f}"
                    if column.key in totals:
                        totals[column.key] += value
                elif column.format == "date" and value:
                    if isinstance(value, str):
                        formatted = value[:10]  # Just date part
                    else:
                        formatted = value.strftime("%Y-%m-%d")
                elif column.format == "number" and isinstance(value, (int, float)):
                    formatted = f"{value:,}"
                    if column.key in totals:
                        totals[column.key] += value
                elif column.format == "percentage" and isinstance(value, (int, float)):
                    formatted = f"{value:.2f}%"
                else:
                    formatted = str(value) if value else ""

                row.append(formatted)
            table_data.append(row)

        # Totals row
        if self.config.include_totals and totals:
            totals_row = []
            for col in self.config.columns:
                if col.key in totals:
                    if col.format == "currency":
                        totals_row.append(
                            f"{self.config.currency_symbol}{totals[col.key]:,.2f}"
                        )
                    else:
                        totals_row.append(f"{totals[col.key]:,}")
                elif len(totals_row) == 0:
                    totals_row.append("Total")
                else:
                    totals_row.append("")
            table_data.append(totals_row)

        # Create table
        col_widths = [col.width * 6 for col in self.config.columns]  # Convert to points
        table = Table(table_data, colWidths=col_widths)

        # Table style
        style = TableStyle(
            [
                # Header
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                # Body
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                # Alternating rows
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F2F2F2")],
                ),
            ]
        )

        # Totals styling
        if self.config.include_totals and totals:
            style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
            style.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2E2E2"))

        table.setStyle(style)
        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)

        return output.getvalue()


# =============================================================================
# CSV Export
# =============================================================================


class CSVExporter:
    """
    P2.56: Simple CSV exporter.
    """

    def __init__(self, config: ExportConfig):
        self.config = config

    def export(self, data: List[Dict[str, Any]]) -> bytes:
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries to export

        Returns:
            CSV file as bytes (UTF-8 with BOM)
        """
        output = io.StringIO()

        # Write headers
        headers = [col.label for col in self.config.columns]
        writer = csv.writer(output)
        writer.writerow(headers)

        # Write data
        for row_data in data[:MAX_EXPORT_ROWS]:
            row = [row_data.get(col.key, "") for col in self.config.columns]
            writer.writerow(row)

        # Convert to bytes with BOM for Excel compatibility
        content = output.getvalue()
        return ("\ufeff" + content).encode("utf-8")


# =============================================================================
# JSON Export
# =============================================================================


class JSONExporter:
    """
    P2.56: JSON exporter with metadata.
    """

    def __init__(self, config: ExportConfig):
        self.config = config

    def export(self, data: List[Dict[str, Any]]) -> bytes:
        """
        Export data to JSON format.

        Args:
            data: List of dictionaries to export

        Returns:
            JSON file as bytes
        """
        export_data = {
            "metadata": {
                "title": self.config.title,
                "generated_at": datetime.now().isoformat(),
                "total_records": len(data),
                "columns": [
                    {"key": col.key, "label": col.label, "format": col.format}
                    for col in self.config.columns
                ],
            },
            "data": data[:MAX_EXPORT_ROWS],
        }

        return json.dumps(
            export_data, ensure_ascii=False, indent=2, default=str
        ).encode("utf-8")


# =============================================================================
# Export Factory
# =============================================================================


def create_exporter(
    format: str, config: ExportConfig
) -> Union[ExcelExporter, PDFExporter, CSVExporter, JSONExporter]:
    """
    Create an exporter based on format.

    Args:
        format: Export format ('xlsx', 'pdf', 'csv', 'json')
        config: Export configuration

    Returns:
        Appropriate exporter instance
    """
    exporters = {
        "xlsx": ExcelExporter,
        "excel": ExcelExporter,
        "pdf": PDFExporter,
        "csv": CSVExporter,
        "json": JSONExporter,
    }

    exporter_class = exporters.get(format.lower())
    if not exporter_class:
        raise ValueError(f"Unknown export format: {format}")

    return exporter_class(config)


def get_content_type(format: str) -> str:
    """Get MIME type for format."""
    content_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
        "csv": "text/csv",
        "json": "application/json",
    }
    return content_types.get(format.lower(), "application/octet-stream")


def get_file_extension(format: str) -> str:
    """Get file extension for format."""
    extensions = {
        "xlsx": ".xlsx",
        "excel": ".xlsx",
        "pdf": ".pdf",
        "csv": ".csv",
        "json": ".json",
    }
    return extensions.get(format.lower(), "")


# =============================================================================
# Predefined Export Configs
# =============================================================================

PRODUCTS_EXPORT_CONFIG = ExportConfig(
    title="Products Report",
    columns=[
        ExportColumn("id", "ID", 8, "number"),
        ExportColumn("name", "Product Name", 30),
        ExportColumn("sku", "SKU", 15),
        ExportColumn("barcode", "Barcode", 15),
        ExportColumn("price", "Price", 12, "currency"),
        ExportColumn("cost", "Cost", 12, "currency"),
        ExportColumn("quantity", "Quantity", 10, "number"),
        ExportColumn("category_name", "Category", 20),
    ],
    include_totals=True,
    totals_columns=["quantity"],
)

INVOICES_EXPORT_CONFIG = ExportConfig(
    title="Invoices Report",
    columns=[
        ExportColumn("id", "ID", 8, "number"),
        ExportColumn("invoice_number", "Invoice #", 15),
        ExportColumn("type", "Type", 12),
        ExportColumn("customer_name", "Customer", 25),
        ExportColumn("subtotal", "Subtotal", 12, "currency"),
        ExportColumn("discount", "Discount", 10, "currency"),
        ExportColumn("tax", "Tax", 10, "currency"),
        ExportColumn("total", "Total", 12, "currency"),
        ExportColumn("status", "Status", 12),
        ExportColumn("created_at", "Date", 12, "date"),
    ],
    include_totals=True,
    totals_columns=["subtotal", "discount", "tax", "total"],
)

CUSTOMERS_EXPORT_CONFIG = ExportConfig(
    title="Customers Report",
    columns=[
        ExportColumn("id", "ID", 8, "number"),
        ExportColumn("name", "Name", 25),
        ExportColumn("email", "Email", 25),
        ExportColumn("phone", "Phone", 15),
        ExportColumn("balance", "Balance", 12, "currency"),
        ExportColumn("credit_limit", "Credit Limit", 12, "currency"),
    ],
    include_totals=True,
    totals_columns=["balance"],
)


__all__ = [
    "ExcelExporter",
    "PDFExporter",
    "CSVExporter",
    "JSONExporter",
    "ExportConfig",
    "ExportColumn",
    "create_exporter",
    "get_content_type",
    "get_file_extension",
    "PRODUCTS_EXPORT_CONFIG",
    "INVOICES_EXPORT_CONFIG",
    "CUSTOMERS_EXPORT_CONFIG",
]
